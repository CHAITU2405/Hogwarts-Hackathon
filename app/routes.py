from flask import Blueprint, request, jsonify, send_from_directory, current_app, make_response, session, send_file
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
import os
import shutil
from pathlib import Path
from app.models import db, Team, Member, ProblemStatement, AdminSettings, Admin, TeamLogin, Review, Sponsor
from app.config import Config
from datetime import datetime
import io
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

api_bp = Blueprint('api', __name__, url_prefix='/api')

# Add CORS headers
@api_bp.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    if origin:
        response.headers.add('Access-Control-Allow-Origin', origin)
        response.headers.add('Access-Control-Allow-Credentials', 'true')
    else:
        response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,X-Admin-Auth')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# Handle CORS preflight requests
@api_bp.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = make_response()
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,Authorization")
        response.headers.add('Access-Control-Allow-Methods', "GET,PUT,POST,DELETE,OPTIONS")
        return response

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS

def send_credentials_email(receiver_email, team_name, username, password, team_lead_name):
    """Send login credentials to team lead via email using SMTP with multiple connection methods"""
    import socket
    import time
    import ssl
    
    # Get email configuration from Config
    sender_email = Config.SENDER_EMAIL
    sender_password = Config.SENDER_PASSWORD
    smtp_server = Config.SMTP_SERVER
    smtp_port = Config.SMTP_PORT
    
    # Check if email credentials are configured
    if not sender_email or not sender_password:
        print("Email credentials not configured - cannot send email")
        return False
    
    # Create message
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg["Subject"] = f"Welcome to Hogwarts Hackathon - Your Team Login Credentials"
    
    # Email body
    body = f"""Dear {team_lead_name},

Congratulations! Your team "{team_name}" has been approved for the Hogwarts Hackathon.

Your login credentials are as follows:

Username: {username}
Password: {password}

Please keep these credentials secure and use them to log in to the hackathon portal.

Stay Connected:
- Join our WhatsApp Group: https://chat.whatsapp.com/EFpuxL8qmyY9mpfio9Xnw1?mode=hqrc
- Follow us on Instagram: https://www.instagram.com/hogwarts.hackathon?igsh=MWtkZmtiaTVvNTByMQ==

Important Notes:
- You can log in using these credentials on the login page
- Do not share these credentials with anyone outside your team
- Join the WhatsApp group and follow our Instagram for updates, announcements, and to connect with fellow participants
- If you have any issues, please contact the organizers

We look forward to seeing your magical creations!

Best regards,
Hogwarts Hackathon Team"""
    
    msg.attach(MIMEText(body, "plain"))
    
    # Try multiple connection methods for Render compatibility
    # Method 1: TLS on port 587 (standard)
    # Method 2: SSL on port 465 (alternative)
    connection_methods = [
        {'port': smtp_port, 'use_ssl': False, 'use_tls': True, 'name': f'TLS on port {smtp_port}'},
        {'port': 465, 'use_ssl': True, 'use_tls': False, 'name': 'SSL on port 465'},
        {'port': 587, 'use_ssl': False, 'use_tls': True, 'name': 'TLS on port 587'},
    ]
    
    # Remove duplicates if port is already 587 or 465
    seen_ports = set()
    unique_methods = []
    for method in connection_methods:
        if method['port'] not in seen_ports:
            seen_ports.add(method['port'])
            unique_methods.append(method)
    connection_methods = unique_methods
    
    # Retry logic - try each method up to 2 times
    max_retries_per_method = 2
    
    for method in connection_methods:
        for attempt in range(1, max_retries_per_method + 1):
            server = None
            try:
                print(f"Attempting {method['name']} (attempt {attempt}/{max_retries_per_method}) to send email to {receiver_email}")
                
                # Set socket timeout
                socket.setdefaulttimeout(20)  # 20 second timeout for SMTP operations
                
                # Create SMTP connection based on method
                if method['use_ssl']:
                    # SSL connection (port 465)
                    context = ssl.create_default_context()
                    server = smtplib.SMTP_SSL(smtp_server, method['port'], timeout=20, context=context)
                else:
                    # Regular SMTP connection
                    server = smtplib.SMTP(smtp_server, method['port'], timeout=20)
                    if method['use_tls']:
                        # Start TLS encryption
                        server.starttls(context=ssl.create_default_context())
                
                # Login
                server.login(sender_email, sender_password)
                
                # Send email
                server.sendmail(sender_email, [receiver_email], msg.as_string())
                
                # Close connection
                server.quit()
                server = None
                
                print(f"✓ Email sent successfully to {receiver_email} using {method['name']}")
                return True
                
            except smtplib.SMTPAuthenticationError as e:
                error_msg = f"SMTP Authentication Error: {e}"
                print(error_msg)
                print("Check if SENDER_EMAIL and SENDER_PASSWORD are correct")
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                # Don't retry authentication errors - try next method
                break
                
            except (smtplib.SMTPConnectError, ConnectionRefusedError, OSError) as e:
                error_msg = f"SMTP Connection Error ({method['name']}, attempt {attempt}): {e}"
                print(error_msg)
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                if attempt < max_retries_per_method:
                    print(f"Retrying {method['name']} in 2 seconds...")
                    time.sleep(2)
                else:
                    print(f"Failed with {method['name']}, trying next method...")
                    break
                    
            except socket.timeout as e:
                error_msg = f"SMTP Timeout Error ({method['name']}, attempt {attempt}): {e}"
                print(error_msg)
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                if attempt < max_retries_per_method:
                    print(f"Retrying {method['name']} in 2 seconds...")
                    time.sleep(2)
                else:
                    print(f"Timeout with {method['name']}, trying next method...")
                    break
                    
            except smtplib.SMTPException as e:
                error_msg = f"SMTP Error ({method['name']}, attempt {attempt}): {e}"
                print(error_msg)
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                if attempt < max_retries_per_method:
                    print(f"Retrying {method['name']} in 2 seconds...")
                    time.sleep(2)
                else:
                    print(f"Failed with {method['name']}, trying next method...")
                    break
                    
            except Exception as e:
                import traceback
                error_msg = f"Unexpected error ({method['name']}, attempt {attempt}): {e}"
                print(error_msg)
                print(f"Traceback: {traceback.format_exc()}")
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                if attempt < max_retries_per_method:
                    print(f"Retrying {method['name']} in 2 seconds...")
                    time.sleep(2)
                else:
                    print(f"Failed with {method['name']}, trying next method...")
                    break
    
    print(f"✗ Failed to send email to {receiver_email} after trying all connection methods")
    return False

@api_bp.route('/register', methods=['POST'])
def register_team():
    try:
        # Check if registration is enabled
        setting = AdminSettings.query.filter_by(key='registration_enabled').first()
        if not setting or setting.value.lower() != 'true':
            return jsonify({'error': 'Registrations are currently closed'}), 403
        
        # Get form data
        team_name = request.form.get('team_name', '').strip()
        house = request.form.get('house', '').strip()
        team_size = request.form.get('team_size', type=int)
        utr_transaction_id = request.form.get('utr_transaction_id', '').strip()
        
        # Validate required fields
        if not team_name:
            return jsonify({'error': 'Team name is required'}), 400
        
        # Normalize team name (trim and remove extra spaces)
        team_name = ' '.join(team_name.split())
        if not house:
            return jsonify({'error': 'House selection is required'}), 400
        if not team_size or team_size < 1 or team_size > 4:
            return jsonify({'error': 'Team size must be between 1 and 4'}), 400
        if not utr_transaction_id:
            return jsonify({'error': 'UTR/Transaction ID is required'}), 400
        
        # Check if team name already exists (case-insensitive)
        existing_team = Team.query.filter(
            db.func.lower(Team.team_name) == db.func.lower(team_name)
        ).first()
        if existing_team:
            return jsonify({'error': 'Team name already exists'}), 400
        
        # Handle file upload
        payment_proof_path = None
        if 'payment_proof' in request.files:
            file = request.files['payment_proof']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # Create unique filename with timestamp
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                filepath = Path(Config.UPLOAD_FOLDER) / filename
                file.save(str(filepath))
                payment_proof_path = f"uploads/{filename}"
        
        # Create team with pending approval status
        team = Team(
            team_name=team_name,
            house=house.capitalize(),
            team_size=team_size,
            utr_transaction_id=utr_transaction_id,
            payment_proof_path=payment_proof_path,
            approval_status='pending'  # Explicitly set to pending
        )
        db.session.add(team)
        db.session.flush()  # Get team ID
        
        # Get member data from form
        members_data = []
        for i in range(1, team_size + 1):
            name = request.form.get(f'member_{i}_name', '').strip()
            email = request.form.get(f'member_{i}_email', '').strip()
            phone = request.form.get(f'member_{i}_phone', '').strip()
            college_name = request.form.get(f'member_{i}_college', '').strip()
            
            if not name or not email or not phone or not college_name:
                db.session.rollback()
                return jsonify({'error': f'All fields including college name are required for member {i}'}), 400
            
            # Check for duplicate emails
            existing_member = Member.query.filter_by(email=email).first()
            if existing_member:
                db.session.rollback()
                return jsonify({'error': f'Email {email} is already registered'}), 400
            
            member = Member(
                team_id=team.id,
                name=name,
                email=email,
                phone=phone,
                college_name=college_name,
                is_leader=(i == 1),
                member_order=i
            )
            db.session.add(member)
            members_data.append(member)
        
        # Commit all changes to database - this ensures data is stored
        try:
            db.session.commit()
        except Exception as commit_error:
            db.session.rollback()
            return jsonify({'error': f'Failed to save data: {str(commit_error)}'}), 500
        
        # Verify data was saved by querying it back
        saved_team = Team.query.get(team.id)
        if not saved_team:
            return jsonify({'error': 'Registration failed - data not saved'}), 500
        
        # Only return success after confirming data is in database
        response_data = {
            'success': True,
            'message': 'Team registered successfully',
            'team_id': team.id,
            'team': team.to_dict()
        }
        response = jsonify(response_data)
        response.status_code = 201
        return response
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/teams', methods=['GET'])
def get_teams():
    try:
        # Check if teams page is enabled
        setting = AdminSettings.query.filter_by(key='teams_enabled').first()
        if not setting or setting.value.lower() != 'true':
            return jsonify({
                'success': False,
                'error': 'Teams page is currently disabled',
                'enabled': False
            }), 403
        
        house_filter = request.args.get('house', '').strip()
        search_term = request.args.get('search', '').strip()
        
        # Build query
        query = Team.query
        
        # Apply house filter
        if house_filter:
            query = query.filter(Team.house == house_filter)
        
        # Apply search filter
        if search_term:
            query = query.filter(Team.team_name.ilike(f'%{search_term}%'))
        
        teams = query.order_by(Team.registered_at.desc()).all()
        
        # Safely serialize teams
        teams_data = []
        for team in teams:
            try:
                # Use to_dict_summary which has built-in safety
                teams_data.append(team.to_dict_summary())
            except Exception as e:
                import traceback
                print(f"Error serializing team {team.id}: {e}")
                print(traceback.format_exc())
                # Fallback to basic data if serialization fails
                try:
                    # Get members safely
                    members_list = []
                    try:
                        if hasattr(team, 'members') and team.members:
                            members_list = [m.name for m in team.members]
                    except:
                        pass
                    
                    # Get college name from team members (first member's college)
                    college = ''
                    try:
                        if team.members and len(team.members) > 0:
                            first_member = team.members[0]
                            college = getattr(first_member, 'college_name', None) or ''
                    except:
                        pass
                    
                    teams_data.append({
                        'id': team.id,
                        'name': team.team_name,
                        'house': team.house,
                        'members': members_list,
                        'projectUrl': '',
                        'college': college,
                        'description': f'A brave team from {team.house} house',
                        'approval_status': team.approval_status
                    })
                except Exception as e2:
                    import traceback
                    print(f"Error in fallback serialization for team {team.id}: {e2}")
                    print(traceback.format_exc())
                    continue
        
        return jsonify({
            'success': True,
            'teams': teams_data
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in get_teams: {error_trace}")
        return jsonify({'error': str(e), 'trace': error_trace}), 500

@api_bp.route('/teams/<int:team_id>', methods=['GET'])
def get_team(team_id):
    try:
        team = Team.query.get_or_404(team_id)
        return jsonify({
            'success': True,
            'team': team.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/teams/update-repo', methods=['POST'])
def update_team_repo():
    """Update git repository URL for a team"""
    try:
        data = request.get_json()
        team_id = data.get('team_id')
        git_repo_url = data.get('git_repo_url', '').strip()
        
        if not team_id:
            return jsonify({'error': 'Team ID is required'}), 400
        
        team = Team.query.get_or_404(team_id)
        
        # Validate URL format
        if git_repo_url:
            try:
                from urllib.parse import urlparse
                parsed = urlparse(git_repo_url)
                if not parsed.scheme or not parsed.netloc:
                    return jsonify({'error': 'Invalid URL format. Please provide a valid URL (e.g., https://github.com/username/repo)'}), 400
            except Exception:
                return jsonify({'error': 'Invalid URL format. Please provide a valid URL'}), 400
        
        # Update git repo URL
        team.git_repo_url = git_repo_url if git_repo_url else None
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Repository URL updated successfully',
            'team': team.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/uploads/<path:filepath>', methods=['GET'])
def uploaded_file(filepath):
    """Serve uploaded files including files in subdirectories (e.g., uploads/sponsors/logo.jpg)"""
    try:
        upload_folder = Config.UPLOAD_FOLDER
        file_path = upload_folder / filepath
        
        # Security check: ensure file is within upload folder
        try:
            file_path.resolve().relative_to(upload_folder.resolve())
        except ValueError:
            return jsonify({'error': 'Invalid file path'}), 403
        
        if file_path.exists() and file_path.is_file():
            return send_file(str(file_path))
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============ ADMIN ROUTES ============

@api_bp.route('/admin/pending-teams', methods=['GET'])
def get_pending_teams():
    """Get all teams with pending approval status"""
    try:
        teams = Team.query.filter_by(approval_status='pending').order_by(Team.registered_at.desc()).all()
        return jsonify({
            'success': True,
            'teams': [team.to_dict() for team in teams]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/approve-team/<int:team_id>', methods=['POST'])
def approve_team(team_id):
    """Approve a team - sets status to approved and creates login credentials"""
    try:
        team = Team.query.get_or_404(team_id)
        
        # Get team lead (first member or member with is_leader=True)
        team_lead = Member.query.filter_by(team_id=team_id, is_leader=True).first()
        if not team_lead:
            # Fallback to first member if no leader marked
            team_lead = Member.query.filter_by(team_id=team_id).order_by(Member.member_order).first()
        
        if not team_lead:
            return jsonify({'error': 'Team lead not found'}), 400
        
        # Check if login credentials already exist
        existing_login = TeamLogin.query.filter_by(team_id=team_id).first()
        
        if not existing_login:
            # Create login credentials for team lead
            # Username: team lead name, Password: UTR
            team_login = TeamLogin(
                team_id=team.id,
                username=team_lead.name,
                password=team.utr_transaction_id,
                house=team.house
            )
            db.session.add(team_login)
            login_password = team.utr_transaction_id
        else:
            # Update existing login to match current UTR (in case UTR was corrected)
            existing_login.password = team.utr_transaction_id
            existing_login.username = team_lead.name
            existing_login.house = team.house
            login_password = team.utr_transaction_id
        
        # Update team size based on current member count (in case members were removed)
        current_member_count = Member.query.filter_by(team_id=team_id).count()
        team.team_size = current_member_count
        
        # Update team approval status
        team.approval_status = 'approved'
        db.session.commit()
        
        # Send credentials email to team lead (non-blocking - don't fail if email fails)
        email_sent = False
        email_error = None
        try:
            # Check if email credentials are configured
            if Config.SENDER_EMAIL and Config.SENDER_PASSWORD:
                email_sent = send_credentials_email(
                    receiver_email=team_lead.email,
                    team_name=team.team_name,
                    username=team_lead.name,
                    password=login_password,
                    team_lead_name=team_lead.name
                )
                if not email_sent:
                    email_error = "Email sending failed - check server logs for details"
            else:
                print("Email credentials not configured - skipping email send")
                email_error = "Email credentials not configured in environment variables"
        except Exception as e:
            email_error = str(e)
            print(f"Email sending failed (non-critical): {email_error}")
            # Don't fail the approval if email fails
        
        # Build response message
        response_message = f'Team {team.team_name} has been approved. Login credentials created.'
        if email_sent:
            response_message += ' Credentials have been sent to the team lead via email.'
        elif email_error:
            response_message += f' Note: Email could not be sent, but credentials are available.'
        
        # Get team data safely
        try:
            team_dict = team.to_dict()
        except Exception as e:
            print(f"Error serializing team: {e}")
            # Fallback to basic team data
            team_dict = {
                'id': team.id,
                'team_name': team.team_name,
                'house': team.house,
                'approval_status': team.approval_status
            }
        
        return jsonify({
            'success': True,
            'message': response_message,
            'team': team_dict,
            'login': {
                'username': team_lead.name,
                'password': login_password,
                'house': team.house
            },
            'email_sent': email_sent
        }), 200
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in approve_team: {error_trace}")
        return jsonify({'error': str(e), 'trace': error_trace}), 500

@api_bp.route('/admin/reject-team/<int:team_id>', methods=['POST'])
def reject_team(team_id):
    """Reject a team - deletes the team and all related data from the database"""
    print(f"=== REJECT TEAM CALLED: Deleting team {team_id} ===")  # Debug log
    try:
        team = Team.query.get_or_404(team_id)
        team_name = team.team_name  # Store name before deletion
        print(f"Found team: {team_name} (ID: {team_id})")
        
        # Delete related records first (due to foreign key constraints)
        # Delete TeamLogin if it exists
        team_login = TeamLogin.query.filter_by(team_id=team_id).first()
        if team_login:
            print(f"Deleting TeamLogin for team {team_id}")
            db.session.delete(team_login)
        
        # Delete Review if it exists
        review = Review.query.filter_by(team_id=team_id).first()
        if review:
            print(f"Deleting Review for team {team_id}")
            db.session.delete(review)
        
        # Delete all members explicitly (cascade should handle this, but being explicit)
        members = Member.query.filter_by(team_id=team_id).all()
        print(f"Deleting {len(members)} members for team {team_id}")
        for member in members:
            db.session.delete(member)
        
        # Delete the team (DO NOT SET STATUS - DELETE IT)
        print(f"Deleting team {team_id} ({team_name})")
        db.session.delete(team)
        
        # Commit all deletions
        db.session.commit()
        print(f"Successfully deleted team {team_id} from database")
        
        # Verify deletion
        verify_team = Team.query.get(team_id)
        if verify_team:
            print(f"ERROR: Team {team_id} still exists after deletion!")
            return jsonify({'error': 'Team still exists after deletion attempt'}), 500
        
        return jsonify({
            'success': True,
            'message': f'Team {team_name} has been rejected and deleted',
            'deleted': True
        }), 200
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"ERROR deleting team {team_id}: {error_trace}")  # Log to console for debugging
        # Return a more user-friendly error message
        error_msg = str(e)
        if 'foreign key constraint' in error_msg.lower() or 'constraint' in error_msg.lower():
            error_msg = 'Database constraint error. Please ensure all related records are deleted first.'
        elif 'IntegrityError' in str(type(e)):
            error_msg = 'Database integrity error. The team may have related records that prevent deletion.'
        return jsonify({'error': error_msg, 'trace': error_trace}), 500

@api_bp.route('/admin/problem-statements', methods=['GET'])
def get_problem_statements():
    """Get all problem statements"""
    try:
        house_filter = request.args.get('house', '').strip()
        domain_filter = request.args.get('domain', '').strip()
        
        query = ProblemStatement.query
        
        if house_filter:
            # Show statements for the house OR statements available to all (house is null)
            query = query.filter(
                db.or_(
                    ProblemStatement.house == house_filter,
                    ProblemStatement.house.is_(None)
                )
            )
        
        if domain_filter:
            query = query.filter(ProblemStatement.domain == domain_filter)
        
        statements = query.order_by(ProblemStatement.created_at.desc()).all()
        
        # Count teams that selected each statement
        result = []
        for stmt in statements:
            count = Team.query.filter_by(selected_problem_statement_id=stmt.id).count()
            stmt_dict = stmt.to_dict()
            stmt_dict['selected_count'] = count
            result.append(stmt_dict)
        
        return jsonify({
            'success': True,
            'statements': result
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/problem-statements', methods=['POST'])
def add_problem_statement():
    """Add a new problem statement"""
    try:
        data = request.get_json()
        
        # Safely get and strip values, handling None cases
        title = str(data.get('title') or '').strip()
        description = str(data.get('description') or '').strip()
        domain = str(data.get('domain') or '').strip()
        difficulty = str(data.get('difficulty') or '').strip()
        house = str(data.get('house') or '').strip() if data.get('house') else None
        
        if not title or not description or not domain or not difficulty:
            return jsonify({'error': 'All fields are required'}), 400
        
        statement = ProblemStatement(
            title=title,
            description=description,
            domain=domain,
            difficulty=difficulty,
            house=house
        )
        db.session.add(statement)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Problem statement added successfully',
            'statement': statement.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/problem-statements/<int:stmt_id>', methods=['DELETE'])
def delete_problem_statement(stmt_id):
    """Delete a problem statement"""
    try:
        statement = ProblemStatement.query.get_or_404(stmt_id)
        db.session.delete(statement)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Problem statement deleted successfully'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/login-toggle', methods=['GET'])
def get_login_toggle():
    """Get current login toggle status"""
    try:
        setting = AdminSettings.query.filter_by(key='login_enabled').first()
        if not setting:
            # Default to disabled
            setting = AdminSettings(key='login_enabled', value='false')
            db.session.add(setting)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': setting.value.lower() == 'true'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/login-toggle', methods=['POST'])
def toggle_login():
    """Toggle login button visibility"""
    try:
        data = request.get_json()
        enabled = data.get('enabled', False)
        
        setting = AdminSettings.query.filter_by(key='login_enabled').first()
        if not setting:
            setting = AdminSettings(key='login_enabled', value='false')
            db.session.add(setting)
        
        setting.value = 'true' if enabled else 'false'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': enabled,
            'message': f'Login button is now {"enabled" if enabled else "disabled"}'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/registration-toggle', methods=['GET'])
def get_registration_toggle():
    """Get current registration toggle status"""
    try:
        setting = AdminSettings.query.filter_by(key='registration_enabled').first()
        if not setting:
            # Default to disabled
            setting = AdminSettings(key='registration_enabled', value='false')
            db.session.add(setting)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': setting.value.lower() == 'true'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/registration-toggle', methods=['POST'])
def toggle_registration():
    """Toggle registration button visibility"""
    try:
        data = request.get_json()
        enabled = data.get('enabled', False)
        
        setting = AdminSettings.query.filter_by(key='registration_enabled').first()
        if not setting:
            setting = AdminSettings(key='registration_enabled', value='false')
            db.session.add(setting)
        
        setting.value = 'true' if enabled else 'false'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': enabled,
            'message': f'Registration is now {"enabled" if enabled else "disabled"}'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/teams-toggle', methods=['GET'])
def get_teams_toggle():
    """Get current teams toggle status"""
    try:
        setting = AdminSettings.query.filter_by(key='teams_enabled').first()
        if not setting:
            # Default to disabled
            setting = AdminSettings(key='teams_enabled', value='false')
            db.session.add(setting)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': setting.value.lower() == 'true'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/teams-toggle', methods=['POST'])
def toggle_teams():
    """Toggle teams page visibility"""
    try:
        data = request.get_json()
        enabled = data.get('enabled', False)
        
        setting = AdminSettings.query.filter_by(key='teams_enabled').first()
        if not setting:
            setting = AdminSettings(key='teams_enabled', value='false')
            db.session.add(setting)
        
        setting.value = 'true' if enabled else 'false'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'enabled': enabled,
            'message': f'Teams page is now {"enabled" if enabled else "disabled"}'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ============ LOGIN ROUTES ============

@api_bp.route('/login', methods=['POST'])
def login():
    """Team lead login - username is team_name, password is UTR
    Admin login - username is 'Harry Potter', password is 'hogwarts school'"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        is_admin = data.get('is_admin', False)
        
        if not username or not password:
            return jsonify({'error': 'Username and password are required'}), 400
        
        # Check for admin login
        if is_admin or (username.lower() == 'harry potter' and password == 'hogwarts school'):
            # Admin credentials
            if username.lower() == 'harry potter' and password == 'hogwarts school':
                # Set admin session
                session['is_admin'] = True
                session['admin_username'] = 'Harry Potter'
                return jsonify({
                    'success': True,
                    'message': 'Admin login successful',
                    'is_admin': True,
                    'admin': {
                        'username': 'Harry Potter',
                        'name': 'Harry Potter'
                    }
                }), 200
            else:
                return jsonify({'error': 'Invalid admin credentials'}), 401
        
        # Team lead login
        # First check if login credentials exist in TeamLogin table
        team_login = TeamLogin.query.filter(
            db.func.lower(TeamLogin.username) == db.func.lower(username)
        ).first()
        
        if not team_login:
            return jsonify({'error': 'Login credentials not found. Team may not be approved yet.'}), 401
        
        # Verify password (UTR) matches
        if team_login.password != password:
            return jsonify({'error': 'Invalid credentials'}), 401
        
        # Get team details
        team = Team.query.get(team_login.team_id)
        if not team:
            return jsonify({'error': 'Team not found'}), 404
        
        # Double-check team is approved
        if team.approval_status != 'approved':
            return jsonify({'error': 'Team is not approved yet'}), 403
        
        # Return team info (without sensitive data)
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'is_admin': False,
            'team': {
                'id': team.id,
                'team_name': team.team_name,
                'house': team.house,
                'selected_problem_statement_id': team.selected_problem_statement_id
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============ PROBLEM STATEMENT SELECTION ============

@api_bp.route('/admin/problem-statements/<int:stmt_id>/teams', methods=['GET'])
def get_teams_for_statement(stmt_id):
    """Get all teams that selected a specific problem statement"""
    try:
        statement = ProblemStatement.query.get_or_404(stmt_id)
        teams = Team.query.filter_by(selected_problem_statement_id=stmt_id).all()
        
        return jsonify({
            'success': True,
            'teams': [team.to_dict_summary() for team in teams]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/select-problem-statement', methods=['POST'])
def select_problem_statement():
    """Team selects/applies for a problem statement"""
    try:
        data = request.get_json()
        team_id = data.get('team_id')
        problem_statement_id = data.get('problem_statement_id')
        
        if not team_id or not problem_statement_id:
            return jsonify({'error': 'Team ID and Problem Statement ID are required'}), 400
        
        team = Team.query.get_or_404(team_id)
        
        # Check if team is approved
        if team.approval_status != 'approved':
            return jsonify({'error': 'Team is not approved yet'}), 403
        
        # Check if team has already selected a problem statement - prevent any resubmission
        if team.selected_problem_statement_id:
            # Check if trying to select the same statement
            if team.selected_problem_statement_id == problem_statement_id:
                return jsonify({'error': 'You have already applied for this problem statement. Resubmission is not allowed.'}), 400
            # Prevent selecting a different statement - no resubmissions allowed
            return jsonify({'error': 'You have already applied for a problem statement. Resubmission or changing your selection is not allowed.'}), 400
        
        problem_statement = ProblemStatement.query.get_or_404(problem_statement_id)
        
        # Update team's selected problem statement
        team.selected_problem_statement_id = problem_statement_id
        
        # If the problem statement's domain is different from team's house, update team's house
        # Domain values: gryffindor, slytherin, ravenclaw, hufflepuff, muggles
        # Convert domain to capitalized house name for consistency
        problem_house = problem_statement.domain.capitalize() if problem_statement.domain else None
        
        if problem_house and problem_house.lower() != team.house.lower():
            team.house = problem_house
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Problem statement selected successfully',
            'team': team.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/teams', methods=['GET'])
def get_all_teams():
    """Get all approved teams for review marks dropdown"""
    try:
        teams = Team.query.filter_by(approval_status='approved').order_by(Team.team_name).all()
        return jsonify({
            'success': True,
            'teams': [{'id': team.id, 'team_name': team.team_name, 'house': team.house} for team in teams]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/generate-ticket/<int:team_id>', methods=['GET'])
def generate_ticket(team_id):
    """Generate and return a downloadable ticket for the team"""
    try:
        team = Team.query.get_or_404(team_id)
        
        # Check if team is approved
        if team.approval_status != 'approved':
            return jsonify({'error': 'Team is not approved yet'}), 403
        
        # Get team members
        members = Member.query.filter_by(team_id=team_id).order_by(Member.member_order).all()
        
        # Get selected problem statement if any
        problem_statement = None
        if team.selected_problem_statement_id:
            problem_statement = ProblemStatement.query.get(team.selected_problem_statement_id)
        
        # Get house crest image and convert to base64
        house_crests = {
            'Gryffindor': 'gryffindor.png',
            'Slytherin': 'slytherin.png',
            'Ravenclaw': 'ravenclaw.png',
            'Hufflepuff': 'hufflepuff.png',
            'Muggles': 'muggles.png'
        }
        crest_filename = house_crests.get(team.house, 'muggles.png')
        crest_path = Config.BASE_DIR / 'assets' / crest_filename
        
        # Convert image to base64
        crest_base64 = ''
        if crest_path.exists():
            try:
                with open(crest_path, 'rb') as img_file:
                    img_data = img_file.read()
                    crest_base64 = base64.b64encode(img_data).decode('utf-8')
                    crest_base64 = f'data:image/png;base64,{crest_base64}'
            except Exception as e:
                print(f"Error reading crest image: {e}")
                crest_base64 = ''
        
        # Generate HTML ticket in Hogwarts Express style
        ticket_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Hogwarts Express Ticket</title>
    <link href="https://fonts.googleapis.com/css2?family=Cinzel+Decorative:wght@700;900&family=Crimson+Text:ital,wght@0,400;1,400&family=Pinyon+Script&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Crimson Text', serif; 
            background: #8b7355; 
            padding: 40px 20px; 
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }}
        .ticket-wrapper {{
            max-width: 900px;
            width: 100%;
        }}
        .ticket-container {{ 
            background: #f3e9d2;
            background-image: url('https://www.transparenttextures.com/patterns/aged-paper.png');
            border: 4px solid #2c1b18;
            display: flex;
            position: relative;
            box-shadow: 0 10px 40px rgba(0,0,0,0.5);
        }}
        .ticket-stub {{
            width: 35%;
            padding: 30px 20px;
            border-right: 2px dashed #8b7355;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: space-between;
            position: relative;
        }}
        .platform-text {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 1rem;
            color: #2c1b18;
            text-transform: uppercase;
            letter-spacing: 2px;
            margin-bottom: 15px;
        }}
        .platform-number {{
            width: 80px;
            height: 80px;
            border-radius: 50%;
            background: #2c1b18;
            color: #f3e9d2;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.8rem;
            font-weight: bold;
            margin: 20px 0;
            font-family: Arial, sans-serif;
        }}
        .house-crest {{
            width: 120px;
            height: 120px;
            margin: 20px 0;
            background: white;
            border: 2px solid #8b7355;
            border-radius: 8px;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 10px;
        }}
        .house-crest img {{
            max-width: 100%;
            max-height: 100%;
            object-fit: contain;
        }}
        .house-name {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 0.9rem;
            color: #2c1b18;
            text-transform: uppercase;
            font-weight: bold;
            margin-top: 10px;
        }}
        .location-text {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 1rem;
            color: #2c1b18;
            text-transform: uppercase;
            letter-spacing: 2px;
            margin-top: 15px;
        }}
        .ticket-main {{
            width: 65%;
            padding: 40px 30px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }}
        .ticket-title {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 3rem;
            color: #2c1b18;
            text-align: center;
            margin-bottom: 5px;
            font-weight: 900;
            text-shadow: 2px 2px 4px rgba(212, 175, 55, 0.3);
        }}
        .ticket-subtitle {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 2.5rem;
            color: #2c1b18;
            text-align: center;
            margin-bottom: 20px;
            font-weight: 900;
            text-shadow: 2px 2px 4px rgba(212, 175, 55, 0.3);
        }}
        .divider {{
            border-top: 2px double #d4af37;
            margin: 15px 0;
        }}
        .destination {{
            font-family: 'Pinyon Script', cursive;
            font-size: 2rem;
            color: #2c1b18;
            text-align: center;
            margin: 10px 0;
        }}
        .trip-type {{
            font-family: Arial, sans-serif;
            font-size: 0.9rem;
            color: #2c1b18;
            text-align: center;
            text-transform: uppercase;
            letter-spacing: 3px;
            margin: 10px 0;
            font-weight: bold;
        }}
        .team-info {{
            margin: 20px 0;
            padding: 15px;
            background: rgba(255,255,255,0.3);
            border-left: 3px solid #d4af37;
        }}
        .team-name {{
            font-family: 'Cinzel Decorative', serif;
            font-size: 1.3rem;
            color: #8b6914;
            margin-bottom: 10px;
        }}
        .team-details {{
            font-size: 0.95rem;
            color: #2c1b18;
            line-height: 1.6;
        }}
        .members-list {{
            margin: 15px 0;
            font-size: 0.9rem;
            color: #2c1b18;
        }}
        .member-item {{
            margin: 5px 0;
            padding-left: 15px;
        }}
        .rules-text {{
            font-family: Arial, sans-serif;
            font-size: 0.7rem;
            color: #2c1b18;
            text-align: center;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-top: 20px;
            line-height: 1.4;
        }}
        @media print {{
            body {{ padding: 10px; background: white; }}
            .ticket-container {{ box-shadow: none; }}
        }}
    </style>
</head>
<body>
    <div class="ticket-wrapper">
        <div class="ticket-container">
            <!-- Left Stub -->
            <div class="ticket-stub">
                <div class="platform-text">PLATFORM</div>
                <div class="platform-number">9¾</div>
                <div class="house-crest">
                    <img src="{crest_base64}" alt="{team.house} Crest" style="max-width: 100%; max-height: 100%; object-fit: contain;">
                </div>
                <div class="house-name">{team.house.upper()}</div>
                <div class="location-text">LONDON</div>
            </div>
            
            <!-- Main Ticket -->
            <div class="ticket-main">
                <div>
                    <h1 class="ticket-title">HOGWARTS</h1>
                    <h2 class="ticket-subtitle">EXPRESS</h2>
                    <div class="divider"></div>
                    <div class="destination">Sto Hogwarts Hackathon</div>
                    <div class="trip-type">ONE WAY TRIP</div>
                    <div class="divider"></div>
                </div>
                
                <div class="team-info">
                    <div class="team-name">{team.team_name}</div>
                    <div class="team-details">
                        <strong>House:</strong> {team.house} | <strong>Team Size:</strong> {team.team_size} members<br>
                        <strong>Registration:</strong> {team.registered_at.strftime('%B %d, %Y') if team.registered_at else 'N/A'}<br>
                        <strong>UTR:</strong> {team.utr_transaction_id}
                    </div>
                    <div class="members-list">
                        <strong>Team Members:</strong>
"""
        
        for member in members:
            leader_badge = " (Leader)" if member.is_leader else ""
            ticket_html += f"""
                        <div class="member-item">• {member.name}{leader_badge}</div>
"""
        
        ticket_html += """
                    </div>
                </div>
                
                <div class="rules-text">
                    INSTRUCTED TO FOLLOW THE RULES AND REGULATIONS OF HOGWARTS HACKATHON
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""
        
        # Return as downloadable HTML file
        response = make_response(ticket_html)
        response.headers['Content-Type'] = 'text/html'
        response.headers['Content-Disposition'] = f'attachment; filename="Hogwarts_Hackathon_Ticket_{team.team_name.replace(" ", "_")}.html"'
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/statistics', methods=['GET'])
def get_statistics():
    """Get accurate statistics: total members, total teams, and counts by domain"""
    try:
        from sqlalchemy import func
        
        # Total members count (all members across all teams)
        total_members = db.session.query(func.count(Member.id)).scalar() or 0
        
        # Total teams count (all teams, regardless of approval status)
        total_teams = db.session.query(func.count(Team.id)).scalar() or 0
        
        # Count teams by their house (case-insensitive using SQL)
        domain_counts = {}
        houses = ['gryffindor', 'slytherin', 'ravenclaw', 'hufflepuff', 'muggles']
        
        for house in houses:
            # Use case-insensitive comparison with func.lower()
            count = db.session.query(func.count(Team.id)).filter(
                db.func.lower(Team.house) == house
            ).scalar() or 0
            domain_counts[house] = count
        
        return jsonify({
            'success': True,
            'statistics': {
                'total_members': total_members,
                'total_teams': total_teams,
                'by_domain': domain_counts
            }
        }), 200
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in get_statistics: {error_trace}")
        return jsonify({'error': str(e), 'trace': error_trace}), 500

@api_bp.route('/admin/download-database', methods=['GET'])
def download_database():
    """Download the database file (Admin only)"""
    try:
        # Check if admin (simple check - in production, use proper session/auth)
        # For now, we'll allow this endpoint but it should be protected by admin authentication
        # You can add proper admin session checking here
        
        db_path = Config.DATABASE_PATH
        
        if not db_path.exists():
            return jsonify({'error': 'Database file not found'}), 404
        
        # Send the database file
        return send_file(
            str(db_path),
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name='hogwarts_hackathon.db'
        )
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error downloading database: {error_trace}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/upload-database', methods=['POST'])
def upload_database():
    """Upload and overwrite the database file (Admin only)"""
    try:
        # Check if admin is logged in
        is_admin_session = session.get('is_admin', False)
        admin_header = request.headers.get('X-Admin-Auth', '').lower() == 'true'
        
        if not is_admin_session and not admin_header:
            return jsonify({'error': 'Unauthorized. Please log in as admin first.'}), 401
        
        # Check if database file is provided
        if 'database' not in request.files:
            return jsonify({'error': 'No database file provided'}), 400
        
        file = request.files['database']
        
        if not file or not file.filename:
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith('.db'):
            return jsonify({'error': 'Invalid file type. Please upload a .db file'}), 400
        
        db_path = Config.DATABASE_PATH
        
        # Create backup of current database before overwriting (optional safety measure)
        backup_path = None
        if db_path.exists():
            try:
                backup_path = db_path.parent / f'hogwarts_hackathon_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
                shutil.copy2(str(db_path), str(backup_path))
                print(f"Backup created: {backup_path}")
            except Exception as backup_error:
                print(f"Warning: Could not create backup: {backup_error}")
                # Continue anyway - user has been warned
        
        # Close any existing database connections and remove all connections from the engine
        try:
            db.session.close()
            if hasattr(db, 'engine'):
                db.engine.dispose()
        except Exception as close_error:
            print(f"Warning: Error closing database connections: {close_error}")
        
        # Save the uploaded file, overwriting the current database
        file.save(str(db_path))
        
        print(f"Database uploaded and overwritten: {db_path}")
        
        return jsonify({
            'success': True,
            'message': 'Database uploaded and overwritten successfully',
            'backup_created': backup_path.name if backup_path and backup_path.exists() else None
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error uploading database: {error_trace}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/all-teams', methods=['GET'])
def get_all_teams_with_members():
    """Get all approved teams with full member details for management"""
    try:
        # Only get approved teams
        teams = Team.query.filter_by(approval_status='approved').order_by(Team.team_name).all()
        teams_data = []
        for team in teams:
            try:
                teams_data.append(team.to_dict())
            except Exception as e:
                print(f"Error serializing team {team.id}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'teams': teams_data
        }), 200
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in get_all_teams_with_members: {error_trace}")
        return jsonify({'error': str(e), 'trace': error_trace}), 500

@api_bp.route('/admin/teams/<int:team_id>/members', methods=['POST'])
def add_team_member(team_id):
    """Add a new member to a team"""
    try:
        team = Team.query.get_or_404(team_id)
        
        # Check current team size
        current_members = Member.query.filter_by(team_id=team_id).count()
        if current_members >= 4:
            return jsonify({'error': 'Team already has maximum 4 members'}), 400
        
        data = request.get_json()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        college_name = data.get('college_name', '').strip()
        is_leader = data.get('is_leader', False)
        
        if not name or not email or not phone or not college_name:
            return jsonify({'error': 'Name, email, phone, and college name are required'}), 400
        
        # Get next member order
        max_order = db.session.query(db.func.max(Member.member_order)).filter_by(team_id=team_id).scalar() or 0
        member_order = max_order + 1
        
        # Determine if this member should be leader
        # First member is always leader, or if explicitly marked as leader
        will_be_leader = current_members == 0 or is_leader
        
        # If setting as leader, unset other leaders
        if will_be_leader and current_members > 0:
            Member.query.filter_by(team_id=team_id, is_leader=True).update({'is_leader': False})
        
        new_member = Member(
            team_id=team_id,
            name=name,
            email=email,
            phone=phone,
            college_name=college_name,
            is_leader=will_be_leader,
            member_order=member_order
        )
        
        db.session.add(new_member)
        
        # Update team size
        team.team_size = current_members + 1
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Member added successfully',
            'member': new_member.to_dict(),
            'team': team.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/teams/<int:team_id>/members/<int:member_id>', methods=['DELETE'])
def remove_team_member(team_id, member_id):
    """Remove a member from a team"""
    try:
        team = Team.query.get_or_404(team_id)
        member = Member.query.filter_by(id=member_id, team_id=team_id).first_or_404()
        
        # Don't allow removing if it's the only member
        current_members = Member.query.filter_by(team_id=team_id).count()
        if current_members <= 1:
            return jsonify({'error': 'Cannot remove the last member from a team'}), 400
        
        # If removing the leader, assign leadership to the first remaining member
        if member.is_leader:
            remaining_members = Member.query.filter_by(team_id=team_id).filter(Member.id != member_id).order_by(Member.member_order).all()
            if remaining_members:
                remaining_members[0].is_leader = True
        
        # Delete the member
        db.session.delete(member)
        
        # Update team size
        team.team_size = current_members - 1
        
        # Reorder remaining members
        remaining_members = Member.query.filter_by(team_id=team_id).order_by(Member.member_order).all()
        for idx, mem in enumerate(remaining_members, 1):
            mem.member_order = idx
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Member removed successfully',
            'team': team.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/review-marks', methods=['POST'])
def add_review_marks():
    """Add or update review marks for a team - stores all reviews in one row"""
    try:
        import json
        data = request.get_json()
        team_id = data.get('team_id')
        review_number = data.get('review_number')
        marks = data.get('marks')
        feedback = data.get('feedback', '').strip()
        criteria = data.get('criteria', [])  # List of criteria objects
        
        if not team_id or not review_number or marks is None:
            return jsonify({'error': 'Team ID, Review Number, and Marks are required'}), 400
        
        if review_number not in [1, 2, 3]:
            return jsonify({'error': 'Review number must be 1, 2, or 3'}), 400
        
        if marks < 0:
            return jsonify({'error': 'Marks must be non-negative'}), 400
        
        if not feedback:
            return jsonify({'error': 'Feedback is required'}), 400
        
        # Check if team exists and is approved
        team = Team.query.get_or_404(team_id)
        if team.approval_status != 'approved':
            return jsonify({'error': 'Team is not approved'}), 403
        
        # Get or create review row for this team (one row per team)
        review = Review.query.filter_by(team_id=team_id).first()
        
        # Prepare review data as JSON
        review_data = {
            'feedback': feedback,
            'criteria': criteria
        }
        review_data_json = json.dumps(review_data)
        
        if review:
            # Update existing review row
            setattr(review, f'review{review_number}_marks', marks)
            setattr(review, f'review{review_number}_data', review_data_json)
            review.updated_at = datetime.utcnow()
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Review {review_number} updated successfully for team {team.team_name}',
                'review': review.get_review(review_number),
                'is_update': True
            }), 200
        else:
            # Create new review row for this team
            new_review = Review(
                team_id=team_id
            )
            setattr(new_review, f'review{review_number}_marks', marks)
            setattr(new_review, f'review{review_number}_data', review_data_json)
            db.session.add(new_review)
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Review {review_number} added successfully for team {team.team_name}',
                'review': new_review.get_review(review_number),
                'is_update': False
            }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/review-marks/export', methods=['GET'])
def export_review_marks():
    """Export all review marks as Excel file with 3 sheets (Review 1, 2, 3)"""
    try:
        import json
        
        if not OPENPYXL_AVAILABLE:
            # Fallback to JSON if openpyxl not available
            reviews = Review.query.join(Team).filter(Team.approval_status == 'approved').order_by(Team.team_name).all()
            review_data = {1: [], 2: [], 3: []}
            
            for review in reviews:
                for review_num in [1, 2, 3]:
                    review_info = review.get_review(review_num)
                    if review_info and review_info.get('marks', 0) > 0:
                        review_data[review_num].append({
                            'team_id': review.team_id,
                            'team_name': review.team.team_name,
                            'house': review.team.house,
                            'review_number': review_num,
                            'total_marks': review_info.get('marks', 0),
                            'feedback': review_info.get('feedback', ''),
                            'criteria': review_info.get('criteria', [])
                        })
            
            return jsonify({
                'success': True,
                'data': review_data
            }), 200
        
        # Get all reviews with team information (one row per team now)
        reviews = Review.query.join(Team).filter(Team.approval_status == 'approved').order_by(Team.team_name).all()
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each review (1, 2, 3)
        for review_num in [1, 2, 3]:
            # Create sheet for this review
            ws = wb.create_sheet(title=f'Review {review_num}')
            
            # Collect all teams and their criteria for this review
            team_reviews = []
            all_criteria = set()
            
            for review in reviews:
                review_info = review.get_review(review_num)
                if review_info and review_info.get('marks', 0) > 0:
                    criteria_list = review_info.get('criteria', [])
                    # Collect unique criteria names
                    for crit in criteria_list:
                        if crit.get('name'):
                            all_criteria.add(crit.get('name'))
                    
                    team_reviews.append({
                        'team_name': review.team.team_name,
                        'total_marks': review_info.get('marks', 0),
                        'criteria': criteria_list
                    })
            
            criteria_list = sorted(list(all_criteria))
            
            # Create header row: Team Name, [Criteria columns], Total Marks
            header = ['Team Name'] + criteria_list + ['Total Marks']
            ws.append(header)
            
            # Bold header row
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Add data rows
            for team_review in team_reviews:
                row = [team_review['team_name']]
                
                # Add marks for each criterion
                for crit_name in criteria_list:
                    crit = next((c for c in team_review['criteria'] if c.get('name') == crit_name), None)
                    marks = crit.get('marks', 0) if crit else 0
                    row.append(marks)
                
                # Add total marks
                row.append(team_review['total_marks'])
                ws.append(row)
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col in range(1, len(header) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
                    cell_value = str(row[0].value) if row[0].value else ''
                    max_length = max(max_length, len(cell_value))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='review_marks.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Sponsor Management Endpoints
@api_bp.route('/admin/sponsors', methods=['GET'])
def get_sponsors():
    """Get all sponsors"""
    try:
        sponsors = Sponsor.query.order_by(Sponsor.display_order, Sponsor.created_at).all()
        return jsonify({
            'success': True,
            'sponsors': [sponsor.to_dict() for sponsor in sponsors]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/sponsors', methods=['POST'])
def add_sponsor():
    """Add a new sponsor (admin only)"""
    try:
        # Check if admin is logged in - check both session and request header as fallback
        is_admin_session = session.get('is_admin', False)
        # Also check for admin header (for cases where session cookies aren't working)
        admin_header = request.headers.get('X-Admin-Auth', '').lower() == 'true'
        
        if not is_admin_session and not admin_header:
            # Debug: log session info
            print(f"Session check failed. Session keys: {list(session.keys())}, is_admin: {is_admin_session}, header: {admin_header}")
            return jsonify({'error': 'Unauthorized. Please log in as admin first.'}), 401
        
        # Get form data (for file uploads, use form data, not JSON)
        name = request.form.get('name', '').strip()
        redirect_url = request.form.get('redirect_url', '').strip()
        display_order = int(request.form.get('display_order', 0) or 0)
        
        if not name:
            return jsonify({'error': 'Sponsor name is required'}), 400
        
        # Handle file upload - logo is required as file
        logo_path = None
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # Add timestamp to avoid conflicts
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                filename = timestamp + filename
                sponsors_dir = Config.UPLOAD_FOLDER / 'sponsors'
                sponsors_dir.mkdir(parents=True, exist_ok=True)
                file_path = sponsors_dir / filename
                file.save(str(file_path))
                logo_path = f'uploads/sponsors/{filename}'
            else:
                return jsonify({'error': 'Invalid logo file. Please upload a valid image file (PNG, JPG, JPEG, GIF)'}), 400
        else:
            return jsonify({'error': 'Logo image file is required'}), 400
        
        if not logo_path:
            return jsonify({'error': 'Failed to upload logo file'}), 400
        
        sponsor = Sponsor(
            name=name,
            logo_path=logo_path,
            redirect_url=redirect_url if redirect_url else None,
            display_order=display_order
        )
        
        db.session.add(sponsor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sponsor added successfully',
            'sponsor': sponsor.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error adding sponsor: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@api_bp.route('/admin/sponsors/<int:sponsor_id>', methods=['DELETE'])
def delete_sponsor(sponsor_id):
    """Delete a sponsor (admin only)"""
    try:
        # Check if admin is logged in
        if not session.get('is_admin'):
            return jsonify({'error': 'Unauthorized'}), 401
        
        sponsor = Sponsor.query.get_or_404(sponsor_id)
        
        # Delete logo file if it exists
        if sponsor.logo_path:
            # Construct full path - logo_path is stored as 'uploads/sponsors/filename.jpg'
            logo_full_path = Config.BASE_DIR / sponsor.logo_path
            if logo_full_path.exists():
                try:
                    os.remove(str(logo_full_path))
                except Exception as e:
                    print(f"Warning: Could not delete logo file {logo_full_path}: {e}")
                    pass
        
        db.session.delete(sponsor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sponsor deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/sponsors', methods=['GET'])
def get_public_sponsors():
    """Get all sponsors for public display"""
    try:
        sponsors = Sponsor.query.order_by(Sponsor.display_order, Sponsor.created_at).all()
        return jsonify({
            'success': True,
            'sponsors': [sponsor.to_dict() for sponsor in sponsors]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def register_blueprints(app):
    app.register_blueprint(api_bp)
    Config.init_app(app)

